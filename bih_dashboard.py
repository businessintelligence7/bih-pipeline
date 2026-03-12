"""
BIH Marketing AI — Operations Dashboard
========================================
A local web dashboard that shows the real-time status of the
entire BIH Marketing AI system.

WHAT IT SHOWS:
  - Content queue status (posts pending, posted, by platform)
  - Last run times for each script
  - Recent log activity
  - Today's scheduled posts
  - System health indicators

HOW TO RUN:
  python bih_dashboard.py
  Then open: http://localhost:5050

INSTALL (run once):
  pip install flask openpyxl
"""

from flask import Flask, jsonify, render_template_string
from openpyxl import load_workbook
from datetime import datetime, date
import os, json, re

app = Flask(__name__)

# ─────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────
RESOURCES_DIR   = r"C:\Users\ohutchinson\Documents\Omaro\Business\BIH\Resources"
CONTENT_FILE    = os.path.join(RESOURCES_DIR, "BIH_Content_Queue.xlsx")
HEADLINES_FILE  = os.path.join(RESOURCES_DIR, "BIH_Marketing_Headlines.xlsx")
LOGS = {
    "Content Generator": os.path.join(RESOURCES_DIR, "bih_claude_generator_log.txt"),
    "Content Scheduler": os.path.join(RESOURCES_DIR, "bih_content_log.txt"),
    "News Monitor":      os.path.join(RESOURCES_DIR, "bih_news_monitor_log.txt"),
}
SCHEDULE = [
    {"time": "08:45 AM", "script": "Claude Content Generator", "status_key": "Content Generator"},
    {"time": "09:00 AM", "script": "News Monitor",             "status_key": "News Monitor"},
    {"time": "09:15 AM", "script": "Content Scheduler",        "status_key": "Content Scheduler"},
]

# ─────────────────────────────────────────────────────────────
# DATA HELPERS
# ─────────────────────────────────────────────────────────────

def read_queue():
    data = {"total": 0, "pending": 0, "posted": 0, "failed": 0,
            "by_platform": {"LinkedIn": 0, "WhatsApp": 0, "Instagram": 0},
            "today_posts": [], "recent_posts": []}
    if not os.path.exists(CONTENT_FILE):
        return data
    try:
        wb = load_workbook(CONTENT_FILE)
        ws = wb.active
        today = date.today().strftime("%Y-%m-%d")
        rows  = list(ws.iter_rows(min_row=2, values_only=True))
        rows  = [r for r in rows if r and r[0]]
        data["total"] = len(rows)
        for row in rows:
            pid, created, post_date, platform, persona, status, caption, hashtags, notes = (list(row) + [None]*9)[:9]
            status   = status   or "PENDING"
            platform = platform or ""
            if status == "PENDING":   data["pending"] += 1
            elif status == "POSTED":  data["posted"]  += 1
            elif status == "FAILED":  data["failed"]  += 1
            if platform in data["by_platform"]:
                data["by_platform"][platform] += 1
            post = {
                "id": pid, "date": str(post_date or ""), "platform": platform,
                "status": status, "caption": str(caption or "")[:120] + "..." if caption and len(str(caption)) > 120 else str(caption or ""),
                "persona": persona or "",
            }
            if str(post_date or "") == today:
                data["today_posts"].append(post)
        # Last 6 posts
        data["recent_posts"] = [
            {"id": r[0], "date": str(r[2] or ""), "platform": r[3] or "",
             "status": r[5] or "", "caption": str(r[6] or "")[:100] + "..." if r[6] and len(str(r[6])) > 100 else str(r[6] or "")}
            for r in rows[-6:]
        ][::-1]
    except Exception as e:
        data["error"] = str(e)
    return data


def read_log_last(log_path, n=8):
    """Return last n lines of a log file."""
    if not os.path.exists(log_path):
        return []
    try:
        with open(log_path, "r", encoding="utf-8") as f:
            lines = [l.strip() for l in f.readlines() if l.strip()]
        return lines[-n:]
    except:
        return []


def get_last_run(log_path):
    """Get timestamp of last log entry."""
    lines = read_log_last(log_path, 20)
    for line in reversed(lines):
        match = re.match(r"\[(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\]", line)
        if match:
            try:
                dt   = datetime.strptime(match.group(1), "%Y-%m-%d %H:%M:%S")
                diff = datetime.now() - dt
                mins = int(diff.total_seconds() // 60)
                if mins < 60:    return f"{mins}m ago"
                elif mins < 1440: return f"{mins // 60}h ago"
                else:            return f"{mins // 1440}d ago"
            except:
                pass
    return "Never"


def get_system_status():
    statuses = {}
    for name, path in LOGS.items():
        last = get_last_run(path)
        lines = read_log_last(path, 5)
        ok    = any("Done" in l or "OK" in l or "Saved" in l or "posted" in l.lower() for l in lines)
        err   = any("ERROR" in l or "FAIL" in l for l in lines)
        statuses[name] = {
            "last_run": last,
            "health":   "error" if err else ("ok" if ok else "idle"),
            "log":      lines,
        }
    return statuses


def get_headlines():
    if not os.path.exists(HEADLINES_FILE):
        return []
    try:
        wb   = load_workbook(HEADLINES_FILE)
        ws   = wb.active
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
        return [{"title": str(r[0])[:90], "source": str(r[1]) if len(r) > 1 and r[1] else ""} for r in rows[-5:]][::-1]
    except:
        return []


# ─────────────────────────────────────────────────────────────
# API ENDPOINTS
# ─────────────────────────────────────────────────────────────

@app.route("/api/data")
def api_data():
    queue   = read_queue()
    system  = get_system_status()
    headlines = get_headlines()
    return jsonify({
        "queue":      queue,
        "system":     system,
        "headlines":  headlines,
        "timestamp":  datetime.now().strftime("%H:%M:%S"),
        "today":      date.today().strftime("%A, %B %d, %Y"),
    })


# ─────────────────────────────────────────────────────────────
# DASHBOARD HTML
# ─────────────────────────────────────────────────────────────

DASHBOARD_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>BIH Marketing AI — Dashboard</title>
<link href="https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@400;600;700;800&family=Barlow:wght@300;400;500&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
:root {
  --blue:   #1A3FD4;
  --blue2:  #0066FF;
  --gold:   #F5C400;
  --gold2:  #FFD740;
  --bg:     #05080F;
  --bg2:    #080D18;
  --bg3:    #0C1220;
  --border: rgba(26,63,212,0.2);
  --text:   #E8EDF8;
  --muted:  rgba(232,237,248,0.45);
  --green:  #00C896;
  --red:    #FF4757;
  --amber:  #FFB300;
}
* { box-sizing: border-box; margin: 0; padding: 0; }
body {
  font-family: 'Barlow', sans-serif;
  background: var(--bg);
  color: var(--text);
  min-height: 100vh;
  background-image:
    radial-gradient(ellipse 80% 50% at 50% -10%, rgba(26,63,212,0.15) 0%, transparent 60%),
    repeating-linear-gradient(0deg, transparent, transparent 40px, rgba(26,63,212,0.03) 40px, rgba(26,63,212,0.03) 41px),
    repeating-linear-gradient(90deg, transparent, transparent 40px, rgba(26,63,212,0.03) 40px, rgba(26,63,212,0.03) 41px);
}

/* HEADER */
.header {
  display: flex; align-items: center; justify-content: space-between;
  padding: 20px 32px;
  border-bottom: 1px solid var(--border);
  background: rgba(8,13,24,0.9);
  backdrop-filter: blur(10px);
  position: sticky; top: 0; z-index: 100;
}
.header-left { display: flex; align-items: center; gap: 16px; }
.logo-mark {
  width: 38px; height: 38px; border-radius: 8px;
  background: linear-gradient(135deg, var(--blue), var(--blue2));
  display: flex; align-items: center; justify-content: center;
  font-family: 'Barlow Condensed', sans-serif;
  font-weight: 800; font-size: 16px; color: #fff;
  box-shadow: 0 0 20px rgba(26,63,212,0.4);
}
.header-title { font-family: 'Barlow Condensed', sans-serif; font-weight: 700; font-size: 20px; letter-spacing: 0.05em; }
.header-sub { font-family: 'DM Mono', monospace; font-size: 10px; color: var(--muted); letter-spacing: 0.15em; text-transform: uppercase; margin-top: 2px; }
.header-right { display: flex; align-items: center; gap: 20px; }
.live-badge {
  display: flex; align-items: center; gap: 6px;
  font-family: 'DM Mono', monospace; font-size: 10px;
  color: var(--green); letter-spacing: 0.1em;
}
.live-dot { width: 7px; height: 7px; border-radius: 50%; background: var(--green); animation: pulse 2s infinite; }
@keyframes pulse { 0%,100%{opacity:1;transform:scale(1)} 50%{opacity:0.5;transform:scale(1.3)} }
.timestamp { font-family: 'DM Mono', monospace; font-size: 11px; color: var(--muted); }
.refresh-btn {
  padding: 7px 16px; border-radius: 6px; border: 1px solid var(--border);
  background: rgba(26,63,212,0.1); color: var(--text); font-family: 'Barlow Condensed', sans-serif;
  font-size: 13px; font-weight: 600; letter-spacing: 0.05em; cursor: pointer;
  transition: all 0.2s;
}
.refresh-btn:hover { background: rgba(26,63,212,0.25); border-color: rgba(26,63,212,0.5); }

/* LAYOUT */
.main { padding: 28px 32px; max-width: 1400px; margin: 0 auto; }
.date-bar { font-family: 'DM Mono', monospace; font-size: 11px; color: var(--gold); letter-spacing: 0.15em; text-transform: uppercase; margin-bottom: 24px; }

.grid-4 { display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; margin-bottom: 24px; }
.grid-3 { display: grid; grid-template-columns: repeat(3, 1fr); gap: 16px; margin-bottom: 24px; }
.grid-2 { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 24px; }
.grid-21 { display: grid; grid-template-columns: 2fr 1fr; gap: 16px; margin-bottom: 24px; }

/* CARDS */
.card {
  background: var(--bg3);
  border: 1px solid var(--border);
  border-radius: 12px;
  padding: 20px;
  position: relative;
  overflow: hidden;
}
.card::before {
  content: '';
  position: absolute; top: 0; left: 0; right: 0; height: 2px;
  background: linear-gradient(90deg, var(--blue), transparent);
}
.card-gold::before { background: linear-gradient(90deg, var(--gold), transparent); }
.card-green::before { background: linear-gradient(90deg, var(--green), transparent); }

.card-label {
  font-family: 'DM Mono', monospace; font-size: 9px; letter-spacing: 0.2em;
  text-transform: uppercase; color: var(--muted); margin-bottom: 10px;
}
.card-value {
  font-family: 'Barlow Condensed', sans-serif; font-size: 42px;
  font-weight: 800; line-height: 1; color: var(--text);
}
.card-sub { font-size: 12px; color: var(--muted); margin-top: 6px; font-weight: 300; }

/* STAT ACCENT */
.accent-blue { color: var(--blue2); }
.accent-gold { color: var(--gold); }
.accent-green { color: var(--green); }
.accent-red { color: var(--red); }
.accent-amber { color: var(--amber); }

/* SECTION TITLE */
.section-title {
  font-family: 'Barlow Condensed', sans-serif; font-weight: 700;
  font-size: 13px; letter-spacing: 0.12em; text-transform: uppercase;
  color: var(--muted); margin-bottom: 14px; display: flex; align-items: center; gap: 8px;
}
.section-title::after { content: ''; flex: 1; height: 1px; background: var(--border); }

/* SYSTEM CARDS */
.sys-card {
  background: var(--bg3); border: 1px solid var(--border); border-radius: 10px;
  padding: 16px 20px; display: flex; flex-direction: column; gap: 8px;
}
.sys-name { font-family: 'Barlow Condensed', sans-serif; font-weight: 700; font-size: 15px; letter-spacing: 0.05em; }
.sys-meta { display: flex; align-items: center; justify-content: space-between; }
.sys-last { font-family: 'DM Mono', monospace; font-size: 10px; color: var(--muted); }
.health-badge {
  font-family: 'DM Mono', monospace; font-size: 9px; letter-spacing: 0.1em;
  text-transform: uppercase; padding: 3px 9px; border-radius: 20px; font-weight: 500;
}
.health-ok    { background: rgba(0,200,150,0.12); color: var(--green); border: 1px solid rgba(0,200,150,0.25); }
.health-error { background: rgba(255,71,87,0.12);  color: var(--red);   border: 1px solid rgba(255,71,87,0.25); }
.health-idle  { background: rgba(255,179,0,0.12);  color: var(--amber); border: 1px solid rgba(255,179,0,0.25); }

/* PLATFORM BAR */
.platform-row { display: flex; align-items: center; gap: 12px; margin-bottom: 10px; }
.platform-label { font-family: 'Barlow Condensed', sans-serif; font-size: 13px; font-weight: 600; width: 90px; letter-spacing: 0.04em; }
.platform-bar-wrap { flex: 1; height: 6px; background: rgba(255,255,255,0.06); border-radius: 3px; overflow: hidden; }
.platform-bar-fill { height: 100%; border-radius: 3px; transition: width 0.8s ease; }
.bar-li   { background: linear-gradient(90deg, var(--blue), var(--blue2)); }
.bar-wa   { background: linear-gradient(90deg, #00C896, #00E5B0); }
.bar-ig   { background: linear-gradient(90deg, #E1306C, #F77737); }
.platform-count { font-family: 'DM Mono', monospace; font-size: 11px; color: var(--muted); width: 28px; text-align: right; }

/* POST LIST */
.post-item {
  padding: 12px 14px; border-radius: 8px;
  background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.06);
  margin-bottom: 8px; transition: border-color 0.2s;
}
.post-item:hover { border-color: rgba(26,63,212,0.3); }
.post-top { display: flex; align-items: center; gap: 8px; margin-bottom: 6px; }
.post-platform-tag {
  font-family: 'DM Mono', monospace; font-size: 9px; letter-spacing: 0.1em;
  text-transform: uppercase; padding: 2px 8px; border-radius: 20px;
}
.tag-li { background: rgba(26,63,212,0.2); color: #6B9FFF; }
.tag-wa { background: rgba(0,200,150,0.15); color: var(--green); }
.tag-ig { background: rgba(225,48,108,0.15); color: #FF7DAA; }
.post-status-tag {
  font-family: 'DM Mono', monospace; font-size: 9px; padding: 2px 8px;
  border-radius: 20px; margin-left: auto;
}
.status-pending { background: rgba(255,179,0,0.12); color: var(--amber); }
.status-posted  { background: rgba(0,200,150,0.12); color: var(--green); }
.status-failed  { background: rgba(255,71,87,0.12); color: var(--red); }
.post-caption { font-size: 12px; color: var(--muted); line-height: 1.5; font-weight: 300; }
.post-date { font-family: 'DM Mono', monospace; font-size: 9px; color: rgba(255,255,255,0.2); margin-top: 4px; }

/* SCHEDULE */
.sched-item {
  display: flex; align-items: center; gap: 14px;
  padding: 12px 0; border-bottom: 1px solid rgba(255,255,255,0.05);
}
.sched-item:last-child { border-bottom: none; }
.sched-time { font-family: 'DM Mono', monospace; font-size: 12px; color: var(--gold); width: 70px; flex-shrink: 0; }
.sched-name { font-family: 'Barlow Condensed', sans-serif; font-size: 14px; font-weight: 600; letter-spacing: 0.03em; }

/* HEADLINES */
.headline-item {
  padding: 10px 0; border-bottom: 1px solid rgba(255,255,255,0.05);
  font-size: 12px; color: var(--muted); line-height: 1.5;
}
.headline-item:last-child { border-bottom: none; }
.headline-source { font-family: 'DM Mono', monospace; font-size: 9px; color: rgba(255,255,255,0.25); margin-top: 3px; }

/* LOG */
.log-wrap { max-height: 180px; overflow-y: auto; }
.log-line { font-family: 'DM Mono', monospace; font-size: 10px; color: var(--muted); line-height: 1.8; padding: 1px 0; }
.log-line.ok   { color: rgba(0,200,150,0.8); }
.log-line.err  { color: rgba(255,71,87,0.8); }
.log-line.info { color: rgba(100,150,255,0.8); }

/* BUFFER RING */
.buffer-wrap { display: flex; align-items: center; gap: 20px; }
.ring-svg { flex-shrink: 0; }
.ring-info { flex: 1; }
.ring-pct { font-family: 'Barlow Condensed', sans-serif; font-size: 36px; font-weight: 800; }
.ring-label { font-size: 12px; color: var(--muted); font-weight: 300; margin-top: 4px; }

/* LOADING */
.loading { text-align: center; padding: 60px; color: var(--muted); font-family: 'DM Mono', monospace; font-size: 12px; letter-spacing: 0.1em; }

@media (max-width: 900px) {
  .grid-4 { grid-template-columns: repeat(2, 1fr); }
  .grid-3 { grid-template-columns: 1fr; }
  .grid-2, .grid-21 { grid-template-columns: 1fr; }
  .main { padding: 20px 16px; }
}
</style>
</head>
<body>

<header class="header">
  <div class="header-left">
    <div class="logo-mark">BIH</div>
    <div>
      <div class="header-title">Marketing AI Dashboard</div>
      <div class="header-sub">Business Intelligence Holdings</div>
    </div>
  </div>
  <div class="header-right">
    <div class="live-badge"><div class="live-dot"></div>LIVE</div>
    <div class="timestamp" id="clock">--:--:--</div>
    <button class="refresh-btn" onclick="loadData()">↻ Refresh</button>
  </div>
</header>

<div class="main" id="main">
  <div class="loading">Loading dashboard data...</div>
</div>

<script>
let DATA = null;

function clock() {
  document.getElementById('clock').textContent = new Date().toLocaleTimeString('en-US', {hour:'2-digit',minute:'2-digit',second:'2-digit'});
}
setInterval(clock, 1000); clock();

function platformTag(p) {
  if (!p) return '';
  const cls = p === 'LinkedIn' ? 'tag-li' : p === 'WhatsApp' ? 'tag-wa' : 'tag-ig';
  return `<span class="post-platform-tag ${cls}">${p}</span>`;
}
function statusTag(s) {
  if (!s) return '';
  const cls = s === 'POSTED' ? 'status-posted' : s === 'FAILED' ? 'status-failed' : 'status-pending';
  return `<span class="post-status-tag ${cls}">${s}</span>`;
}
function healthBadge(h) {
  const cls = h === 'ok' ? 'health-ok' : h === 'error' ? 'health-error' : 'health-idle';
  const lbl = h === 'ok' ? 'Running' : h === 'error' ? 'Error' : 'Idle';
  return `<span class="health-badge ${cls}">${lbl}</span>`;
}
function logClass(line) {
  if (/OK|Done|Saved|posted|generated/i.test(line)) return 'ok';
  if (/ERROR|FAIL/i.test(line)) return 'err';
  if (/Starting|Checking|Loading/i.test(line)) return 'info';
  return '';
}
function barWidth(val, total) {
  return total === 0 ? 0 : Math.round((val / total) * 100);
}

function render(d) {
  const q  = d.queue;
  const sys = d.system;
  const total = q.total || 1;
  const pct = Math.round((q.pending / 21) * 100);
  const circumference = 2 * Math.PI * 38;
  const dash = (pct / 100) * circumference;

  // Today posts
  const todayHtml = q.today_posts && q.today_posts.length
    ? q.today_posts.map(p => `
      <div class="post-item">
        <div class="post-top">${platformTag(p.platform)}${statusTag(p.status)}</div>
        <div class="post-caption">${p.caption}</div>
      </div>`).join('')
    : `<div class="post-item"><div class="post-caption" style="color:rgba(255,255,255,0.2)">No posts scheduled for today.</div></div>`;

  // Recent posts
  const recentHtml = q.recent_posts && q.recent_posts.length
    ? q.recent_posts.map(p => `
      <div class="post-item">
        <div class="post-top">${platformTag(p.platform)}${statusTag(p.status)}</div>
        <div class="post-caption">${p.caption}</div>
        <div class="post-date">${p.date}</div>
      </div>`).join('')
    : `<div class="post-item"><div class="post-caption" style="color:rgba(255,255,255,0.2)">No posts in queue yet.</div></div>`;

  // System cards
  const sysHtml = Object.entries(sys).map(([name, info]) => `
    <div class="sys-card">
      <div class="sys-meta">
        <div class="sys-name">${name}</div>
        ${healthBadge(info.health)}
      </div>
      <div class="sys-last">Last run: ${info.last_run}</div>
      <div class="log-wrap" style="margin-top:8px">
        ${(info.log || []).map(l => `<div class="log-line ${logClass(l)}">${l}</div>`).join('') || '<div class="log-line">No log entries yet.</div>'}
      </div>
    </div>`).join('');

  // Headlines
  const hlHtml = d.headlines && d.headlines.length
    ? d.headlines.map(h => `
      <div class="headline-item">
        ${h.title}
        ${h.source ? `<div class="headline-source">${h.source}</div>` : ''}
      </div>`).join('')
    : `<div class="headline-item">No headlines loaded yet. Run the News Monitor.</div>`;

  // Schedule
  const schedHtml = [
    {time:'08:45 AM', name:'Claude Content Generator'},
    {time:'09:00 AM', name:'News Monitor'},
    {time:'09:15 AM', name:'Content Scheduler'},
  ].map(s => `
    <div class="sched-item">
      <div class="sched-time">${s.time}</div>
      <div class="sched-name">${s.name}</div>
    </div>`).join('');

  document.getElementById('main').innerHTML = `
    <div class="date-bar">${d.today} &nbsp;·&nbsp; Last updated ${d.timestamp}</div>

    <!-- KPI ROW -->
    <div class="grid-4">
      <div class="card card-gold">
        <div class="card-label">Posts in Queue</div>
        <div class="card-value accent-gold">${q.pending}</div>
        <div class="card-sub">of ${21} target (7-day buffer)</div>
      </div>
      <div class="card card-green">
        <div class="card-label">Total Posted</div>
        <div class="card-value accent-green">${q.posted}</div>
        <div class="card-sub">all time</div>
      </div>
      <div class="card">
        <div class="card-label">Total Generated</div>
        <div class="card-value accent-blue">${q.total}</div>
        <div class="card-sub">LinkedIn + WhatsApp + Instagram</div>
      </div>
      <div class="card">
        <div class="card-label">Failed Posts</div>
        <div class="card-value ${q.failed > 0 ? 'accent-red' : 'accent-green'}">${q.failed}</div>
        <div class="card-sub">${q.failed > 0 ? 'check scheduler log' : 'all clear'}</div>
      </div>
    </div>

    <!-- BUFFER + PLATFORMS + SCHEDULE -->
    <div class="grid-3">
      <div class="card">
        <div class="section-title">Buffer Health</div>
        <div class="buffer-wrap">
          <svg class="ring-svg" width="90" height="90" viewBox="0 0 90 90">
            <circle cx="45" cy="45" r="38" fill="none" stroke="rgba(255,255,255,0.06)" stroke-width="7"/>
            <circle cx="45" cy="45" r="38" fill="none"
              stroke="${pct >= 70 ? '#00C896' : pct >= 40 ? '#FFB300' : '#FF4757'}"
              stroke-width="7" stroke-linecap="round"
              stroke-dasharray="${dash} ${circumference}"
              transform="rotate(-90 45 45)" style="transition:stroke-dasharray 0.8s ease"/>
            <text x="45" y="50" text-anchor="middle" fill="white"
              style="font-family:'Barlow Condensed',sans-serif;font-size:18px;font-weight:800">${pct}%</text>
          </svg>
          <div class="ring-info">
            <div class="ring-pct ${pct >= 70 ? 'accent-green' : pct >= 40 ? 'accent-amber' : 'accent-red'}">${q.pending}<span style="font-size:18px;font-weight:400;color:var(--muted)"> / 21</span></div>
            <div class="ring-label">posts ready to publish</div>
          </div>
        </div>
      </div>

      <div class="card">
        <div class="section-title">By Platform</div>
        <div class="platform-row">
          <div class="platform-label">LinkedIn</div>
          <div class="platform-bar-wrap"><div class="platform-bar-fill bar-li" style="width:${barWidth(q.by_platform.LinkedIn, total)}%"></div></div>
          <div class="platform-count">${q.by_platform.LinkedIn}</div>
        </div>
        <div class="platform-row">
          <div class="platform-label">WhatsApp</div>
          <div class="platform-bar-wrap"><div class="platform-bar-fill bar-wa" style="width:${barWidth(q.by_platform.WhatsApp, total)}%"></div></div>
          <div class="platform-count">${q.by_platform.WhatsApp}</div>
        </div>
        <div class="platform-row">
          <div class="platform-label">Instagram</div>
          <div class="platform-bar-wrap"><div class="platform-bar-fill bar-ig" style="width:${barWidth(q.by_platform.Instagram, total)}%"></div></div>
          <div class="platform-count">${q.by_platform.Instagram}</div>
        </div>
      </div>

      <div class="card">
        <div class="section-title">Daily Schedule</div>
        ${schedHtml}
      </div>
    </div>

    <!-- SYSTEM STATUS -->
    <div class="section-title">System Status</div>
    <div class="grid-3" style="margin-top:0">${sysHtml}</div>

    <!-- TODAY + RECENT -->
    <div class="grid-2">
      <div class="card">
        <div class="section-title">Today's Posts</div>
        ${todayHtml}
      </div>
      <div class="card">
        <div class="section-title">Recent Queue</div>
        ${recentHtml}
      </div>
    </div>

    <!-- HEADLINES -->
    <div class="card" style="margin-bottom:32px">
      <div class="section-title">Latest Headlines (News Monitor)</div>
      ${hlHtml}
    </div>
  `;
}

async function loadData() {
  try {
    const res  = await fetch('/api/data');
    const data = await res.json();
    DATA = data;
    render(data);
  } catch(e) {
    document.getElementById('main').innerHTML = `
      <div class="loading" style="color:#FF4757">
        Could not load data.<br><small>${e.message}</small>
      </div>`;
  }
}

loadData();
setInterval(loadData, 30000); // auto-refresh every 30 seconds
</script>
</body>
</html>"""

@app.route("/")
def dashboard():
    return render_template_string(DASHBOARD_HTML)


if __name__ == "__main__":
    print("\n" + "="*50)
    print("  BIH Marketing AI Dashboard")
    print("  Open: http://localhost:5050")
    print("="*50 + "\n")
    app.run(host="0.0.0.0", port=5050, debug=False)
