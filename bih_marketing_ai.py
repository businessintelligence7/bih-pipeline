"""
BIH Marketing AI — Unified Pipeline
=====================================
Business Intelligence Holdings — Single script that runs the
entire marketing AI pipeline from news to posting.

PIPELINE (runs 3x daily via Windows Task Scheduler):
  08:45 AM  → Step 1: Scrape Jamaica/Caribbean business headlines
              Step 2: Claude generates 3 posts (LinkedIn + WhatsApp + Instagram)
              Step 3: Gemini generates images/video for each post
  09:15 AM  → Step 4: Post to LinkedIn, Instagram, WhatsApp Status
  12:30 PM  → Step 4: Post midday content (video)
  05:30 PM  → Step 4: Post evening content (static image)

DAILY MEDIA PATTERN:
  Morning  (09:00) → Static image   (Gemini Imagen 3)
  Midday   (12:30) → Short video    (Google Veo 2, 10-15 seconds)
  Evening  (05:30) → Static image   (Gemini Imagen 3)

DIMENSIONS:
  Instagram Reel  → 1080x1920 (9:16)
  WhatsApp Status → 1080x1920 (9:16)
  LinkedIn        → 1080x1350 (4:5)

VISUAL STYLE (rotates daily, all platforms share same style):
  Day 1 → Cinematic Dark Tech
  Day 2 → Bold 3D Concept
  Day 3 → Executive Photorealism

FILE PATHS:
  Script  : C:\\Users\\ohutchinson\\Documents\\Omaro\\Business\\BIH\\Sales & Marketing\\bih_marketing_ai.py
  Resources: C:\\Users\\ohutchinson\\Documents\\Omaro\\Business\\BIH\\Sales & Marketing\\Resources\\

INSTALL (run once):
  pip install feedparser anthropic google-genai openpyxl requests pillow

WINDOWS TASK SCHEDULER (3 tasks):
  Task 1: BIH_Marketing_Morning  → 08:45 AM Mon-Sat  (news + generate + post morning)
  Task 2: BIH_Marketing_Midday   → 12:30 PM Mon-Sat  (post midday video)
  Task 3: BIH_Marketing_Evening  → 05:30 PM Mon-Sat  (post evening image)
"""

import os, re, sys, time, json, requests, feedparser, anthropic
from datetime import datetime, date, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from google import genai
    from google.genai import types
    GEMINI_OK = True
except ImportError:
    GEMINI_OK = False

# ═══════════════════════════════════════════════════════════════
# ✏️  CONFIGURATION — fill all values before first run
# ═══════════════════════════════════════════════════════════════
BASE_DIR       = r"C:\Users\ohutchinson\Documents\Omaro\Business\BIH\Sales & Marketing"
RESOURCES_DIR  = os.path.join(BASE_DIR, "Resources")
IMAGES_DIR     = os.path.join(RESOURCES_DIR, "Images")
VIDEOS_DIR     = os.path.join(RESOURCES_DIR, "Videos")
CONTENT_FILE   = os.path.join(RESOURCES_DIR, "BIH_Content_Queue.xlsx")
HEADLINES_FILE = os.path.join(RESOURCES_DIR, "BIH_Marketing_Headlines.xlsx")
LOG_FILE       = os.path.join(RESOURCES_DIR, "bih_marketing_ai_log.txt")

# API Keys
ANTHROPIC_API_KEY = "YOUR_ANTHROPIC_API_KEY"   # console.anthropic.com
GEMINI_API_KEY    = "YOUR_GEMINI_API_KEY"       # aistudio.google.com

# Instagram API (Meta Graph API)
IG_ACCESS_TOKEN   = "YOUR_IG_ACCESS_TOKEN"      # from Meta Developer App
IG_ACCOUNT_ID     = "YOUR_IG_ACCOUNT_ID"        # 17841xxxxxxxxx
IG_APP_ID         = "YOUR_IG_APP_ID"
IG_APP_SECRET     = "YOUR_IG_APP_SECRET"

# LinkedIn API (fill once LinkedIn Developer App is ready)
LI_ACCESS_TOKEN   = "YOUR_LI_ACCESS_TOKEN"
LI_PERSON_URN     = "YOUR_LI_PERSON_URN"        # urn:li:person:xxxxxxxxx

# General
CALENDLY_LINK     = "https://calendly.com/bih"
GMAIL_ADDRESS     = "businessintelligenceinst@gmail.com"
BUFFER_DAYS       = 7
POSTS_PER_DAY     = 3
BUFFER_TARGET     = BUFFER_DAYS * POSTS_PER_DAY  # 21 posts
REST_DAY          = 6   # Sunday

# ═══════════════════════════════════════════════════════════════
# NEWS SOURCES
# ═══════════════════════════════════════════════════════════════
RSS_FEEDS = [
    {"name": "The Gleaner",      "url": "https://jamaica-gleaner.com/feed",             "region": "Local"},
    {"name": "Jamaica Observer", "url": "https://www.jamaicaobserver.com/feed/",        "region": "Local"},
    {"name": "JIS",              "url": "https://jis.gov.jm/feed/",                    "region": "Local"},
    {"name": "TechCrunch",       "url": "https://techcrunch.com/feed/",                "region": "International"},
    {"name": "Bloomberg Tech",   "url": "https://feeds.bloomberg.com/technology/news.rss", "region": "International"},
    {"name": "Fortune",          "url": "https://fortune.com/feed/",                   "region": "International"},
]
RELEVANT_KEYWORDS = [
    "ai","artificial intelligence","automation","machine learning","data","dashboard",
    "analytics","digital transformation","efficiency","productivity","reporting",
    "business intelligence","finance","economy","jamaica","caribbean","business",
    "revenue","growth","enterprise","operations","cost","innovation","strategy","cloud",
]
MAX_PER_FEED = 5

# ═══════════════════════════════════════════════════════════════
# PLATFORM SPECS
# ═══════════════════════════════════════════════════════════════
PLATFORM_SPECS = {
    "LinkedIn":  {"image_aspect": "4:5",  "image_dims": "1080x1350", "video_aspect": "4:5",  "video_dims": "1080x1350"},
    "Instagram": {"image_aspect": "9:16", "image_dims": "1080x1920", "video_aspect": "9:16", "video_dims": "1080x1920"},
    "WhatsApp":  {"image_aspect": "9:16", "image_dims": "1080x1920", "video_aspect": "9:16", "video_dims": "1080x1920"},
}

# ═══════════════════════════════════════════════════════════════
# DAILY ROTATING VISUAL STYLES
# ═══════════════════════════════════════════════════════════════
DAY_STYLES = [
    {
        "name": "Cinematic Dark Tech",
        "mood": "Moody, cinematic, dramatic. Deep shadows, electric blue and gold accent lighting, atmospheric depth, ultra high production value.",
        "image_scenes": {
            "LinkedIn":  "A lone Caribbean executive silhouetted against floor-to-ceiling windows overlooking Kingston at night, holographic financial dashboards floating before them, electric blue light reflecting off polished floors, rim-lit from behind",
            "Instagram": "Vertical cinematic frame: a sharp Caribbean business professional walking confidently through a rain-soaked Kingston street at night, holographic data HUD glowing around them, neon reflections on wet pavement, dramatic backlight",
            "WhatsApp":  "Vertical 9:16 cinematic frame: a confident Caribbean business professional in a sharp navy suit reviewing a holographic dashboard, deep shadows, electric blue accent light, photorealistic",
        },
        "video_scenes": {
            "LinkedIn":  "Cinematic slow push-in through a darkened Caribbean boardroom at night. Holographic financial dashboards and KPI visualisations materialise above the table one by one, glowing electric blue and gold. Atmospheric, no people. 12 seconds.",
            "Instagram": "Vertical 9:16. A Caribbean executive walks in slow motion through a glass office corridor at night. Holographic data streams flow alongside them like a river of blue and gold light. They stop at a window overlooking Kingston — city lights pulse in sync with the data. 12-15 seconds.",
            "WhatsApp":  "Vertical 9:16. Slow cinematic zoom into a glowing real-time business dashboard on a sleek monitor in a dark office. Data updates live — charts move, KPIs pulse with soft blue light. A hand enters frame and taps the screen confidently. 10-12 seconds.",
        },
        "image_technical": "Shot on RED Monstro, 35mm anamorphic, f/1.4, teal and orange LUT, film grain, lens flares, ultra-detailed, 8K. No text, no logos.",
        "video_technical": "Cinematic colour grade, teal and orange LUT, shallow depth of field, subtle film grain, atmospheric. No text, no logos.",
    },
    {
        "name": "Bold 3D Concept",
        "mood": "Vivid, eye-catching, scroll-stopping. High-end 3D render, vibrant accent colours, creative visual metaphors for business intelligence and AI.",
        "image_scenes": {
            "LinkedIn":  "Isometric 3D render of Kingston Jamaica where every building is connected by glowing golden data streams, a massive holographic brain floating above the city pulsing with blue and gold light, epic scale, studio HDRI lighting",
            "Instagram": "Vertical 9:16 bold 3D render: a giant glowing brain made of gold circuitry fills the top half of the frame, lightning bolts connecting it to a Caribbean cityscape below, vivid electric blue sky, epic scale, Blender Cycles render",
            "WhatsApp":  "Vertical 9:16 bold 3D: a 3D rocket ship built entirely from stacked bar charts and data graphs blasting upward through golden clouds, Caribbean ocean visible below, vibrant blue and gold, cinematic scale",
        },
        "video_scenes": {
            "LinkedIn":  "3D animated flythrough starting in outer space pulling into the Caribbean from above. Golden data connection lines appear between islands, zoom into Kingston. Camera descends through clouds to bird's-eye view where every building pulses with live data streams. Ends orbiting a giant glowing holographic brain above the city. 12-15 seconds.",
            "Instagram": "Vertical 9:16 bold 3D animation. A rocket ship built from data charts launches from a Caribbean island. Charts transform into glowing business metrics as it ascends through gold-lit clouds into space. Vivid, energetic. 10-15 seconds.",
            "WhatsApp":  "Vertical 9:16 3D animation. A glowing golden lock in darkness. A circuit-board key flies in and turns. The lock opens with a burst of blue and gold light revealing a thriving digital Caribbean city inside. Camera pushes slowly through the door. 10-12 seconds.",
        },
        "image_technical": "Blender 3D, Cycles renderer, HDRI lighting, subsurface scattering, chromatic aberration, ultra-detailed PBR textures, 8K. No text, no logos.",
        "video_technical": "3D animation, smooth camera movements, vibrant electric blue and gold palette, cinematic timing. No text, no logos.",
    },
    {
        "name": "Executive Photorealism",
        "mood": "Aspirational, credible, human. High-end corporate photography. Authentic Caribbean business context. Shot for Forbes or Harvard Business Review.",
        "image_scenes": {
            "LinkedIn":  "A diverse group of sharp Caribbean business professionals in a modern Kingston boardroom, genuine energy, reviewing live data on curved screens, golden hour light flooding through windows, editorial photography quality",
            "Instagram": "Vertical 9:16 editorial portrait: a powerful Caribbean businesswoman in a crisp blazer walking through a modern Kingston office, one hand gesturing confidently, natural light, authentic and aspirational, Canon R5 85mm f/1.2",
            "WhatsApp":  "Vertical 9:16 lifestyle: a Caribbean business owner in their thriving Kingston shop holding a tablet showing their business dashboard, warm afternoon light, authentic smile, real and relatable",
        },
        "video_scenes": {
            "LinkedIn":  "Documentary-style 4:5 video. A Caribbean executive walks into a modern Kingston boardroom at golden hour, opens a laptop, and a live dashboard populates on the screen and the room's display simultaneously. Executive nods with quiet confidence. Cinematic handheld, warm grade. 12-15 seconds.",
            "Instagram": "Vertical 9:16 lifestyle video. A Caribbean entrepreneur walks through their thriving business — staff working, customers engaged. They pause, check their phone showing a live dashboard, smile. Camera follows their gaze to Kingston through a window. Ends on slow zoom into the glowing dashboard. 10-15 seconds.",
            "WhatsApp":  "Vertical 9:16. A Caribbean professional opens their laptop at a modern desk. Real-time analytics dashboard loads — charts update live, KPIs glow. They lean back satisfied with coffee, looking out at the city. Warm cinematic grade, genuine. 10-12 seconds.",
        },
        "image_technical": "Canon EOS R5, 85mm f/1.2, natural light, editorial photography, high-end retouching, authentic unposed moments. No text, no logos.",
        "video_technical": "Handheld or slider movement, warm cinematic grade, authentic performances, shallow depth of field. No text, no logos.",
    },
]

# ═══════════════════════════════════════════════════════════════
# CONTENT PLATFORM RULES
# ═══════════════════════════════════════════════════════════════
PLATFORM_CONTENT = {
    "LinkedIn": {
        "persona":   "CEO / Managing Director",
        "max_chars": 3000,
        "style":     "Professional thought leadership. 3-5 short paragraphs. Strong hook as first line. Speaks to CEOs, MDs, CFOs at Caribbean businesses. Ends with question or CTA. Hashtags at bottom.",
        "hashtags":  "#BusinessJamaica #CEOmindset #DigitalTransformation #BIH #BusinessIntelligence #Caribbean",
        "cta":       f"Book a free discovery call: {CALENDLY_LINK}",
    },
    "WhatsApp": {
        "persona":   "Trusted Business Advisor",
        "max_chars": 500,
        "style":     "Conversational, warm and direct. 2-3 short sentences max. Feels like a message from a trusted advisor. No hashtags. Ends with a soft nudge.",
        "hashtags":  "",
        "cta":       "Reply or visit bihja.io",
    },
    "Instagram": {
        "persona":   "Caribbean Business Brand",
        "max_chars": 2200,
        "style":     "Punchy and visual. 2-3 short paragraphs. Hook in first line. Emojis sparingly. Hashtags at end.",
        "hashtags":  "#BIH #JamaicaBusiness #Caribbean #BusinessGrowth #Automation #DataDriven #AIBusiness",
        "cta":       "Link in bio → bihja.io",
    },
}

# ═══════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════

def log(msg):
    ts   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except:
        pass

def ensure_dirs():
    for d in [RESOURCES_DIR, IMAGES_DIR, VIDEOS_DIR]:
        os.makedirs(d, exist_ok=True)

def get_slot():
    h, m  = datetime.now().hour, datetime.now().minute
    total = h * 60 + m
    for name, (sh, sm) in [("morning",(9,0)),("midday",(12,30)),("evening",(17,30))]:
        if abs(total - (sh*60+sm)) <= 20:
            return name
    if h < 11:   return "morning"
    elif h < 15: return "midday"
    else:        return "evening"

def get_day_style(target_date=None):
    d = target_date or date.today()
    return DAY_STYLES[d.toordinal() % len(DAY_STYLES)]

def is_relevant(title, summary):
    text    = (title + " " + summary).lower()
    matches = sum(1 for kw in RELEVANT_KEYWORDS if kw in text)
    return matches >= 1

def ensure_content_file():
    if os.path.exists(CONTENT_FILE): return
    wb = Workbook()
    ws = wb.active
    ws.title = "Content Queue"
    headers = ["ID","Date Created","Post Date","Slot","Platform","Persona","Status","Caption","Hashtags","Media Path","Notes"]
    navy = PatternFill("solid", fgColor="1A3FD4")
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = navy
        cell.font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, w in enumerate([10,18,14,12,14,24,12,80,45,50,30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    wb.save(CONTENT_FILE)

def ensure_headlines_file():
    if os.path.exists(HEADLINES_FILE): return
    wb = Workbook()
    ws = wb.active
    ws.title = "Headlines"
    headers = ["Date","Source","Region","Title","Summary","Score"]
    navy = PatternFill("solid", fgColor="1A3FD4")
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = navy
        cell.font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, w in enumerate([18,20,14,80,120,8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(HEADLINES_FILE)

def get_pending_count():
    if not os.path.exists(CONTENT_FILE): return 0
    try:
        wb = load_workbook(CONTENT_FILE)
        ws = wb.active
        return sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r and r[6] == "PENDING")
    except: return 0

def get_next_id():
    if not os.path.exists(CONTENT_FILE): return 1
    try:
        wb   = load_workbook(CONTENT_FILE)
        ws   = wb.active
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
        return len(rows) + 1
    except: return 1

def load_recent_headlines(n=5):
    if not os.path.exists(HEADLINES_FILE): return []
    try:
        wb   = load_workbook(HEADLINES_FILE)
        ws   = wb.active
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and r[3]]
        return [f"- {r[3]}" + (f": {str(r[4])[:100]}" if r[4] else "") for r in rows[-n:]]
    except: return []

def save_posts(posts):
    ensure_content_file()
    wb   = load_workbook(CONTENT_FILE)
    ws   = wb.active
    nid  = get_next_id()
    thin = Border(
        left=Side(style="thin",color="DDDDDD"), right=Side(style="thin",color="DDDDDD"),
        top=Side(style="thin",color="DDDDDD"),  bottom=Side(style="thin",color="DDDDDD"),
    )
    fills = {
        "LinkedIn":  PatternFill("solid", fgColor="DBEAFE"),
        "WhatsApp":  PatternFill("solid", fgColor="DCFCE7"),
        "Instagram": PatternFill("solid", fgColor="FEE2E2"),
    }
    slot_fills = {
        "morning": PatternFill("solid", fgColor="FFF9C4"),
        "midday":  PatternFill("solid", fgColor="E8F5E9"),
        "evening": PatternFill("solid", fgColor="EDE7F6"),
    }
    for p in posts:
        rn   = ws.max_row + 1
        data = [
            f"BIH-{nid:04d}",
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            p["post_date"], p["slot"], p["platform"], p["persona"],
            "PENDING", p["caption"], p["hashtags"], "", p["notes"],
        ]
        for ci, val in enumerate(data, 1):
            cell = ws.cell(row=rn, column=ci, value=val)
            cell.border    = thin
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=(ci in [8,9]))
            if ci == 4: cell.fill = slot_fills.get(p["slot"], PatternFill())
            if ci == 5: cell.fill = fills.get(p["platform"], PatternFill())
            if ci == 7: cell.font = Font(name="Calibri", size=10, bold=True, color="166534")
        ws.row_dimensions[rn].height = 55
        nid += 1
    wb.save(CONTENT_FILE)
    log(f"Saved {len(posts)} posts to queue")

def update_media_path(post_id, media_path, media_type):
    if not os.path.exists(CONTENT_FILE): return
    try:
        wb = load_workbook(CONTENT_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value or "") == post_id:
                row[9].value = f"{media_type}: {media_path}"
        wb.save(CONTENT_FILE)
    except Exception as e:
        log(f"Warning: could not update media path — {e}")

def get_pending_posts_for_slot(slot):
    posts = []
    if not os.path.exists(CONTENT_FILE): return posts
    try:
        wb    = load_workbook(CONTENT_FILE)
        ws    = wb.active
        today = date.today().strftime("%Y-%m-%d")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            pid, created, post_date, pslot, platform, persona, status, caption, hashtags, media, notes = (list(row)+[None]*11)[:11]
            if str(post_date or "")[:10] == today and str(pslot or "") == slot and str(status or "") == "PENDING":
                posts.append({
                    "id": str(pid or ""), "slot": str(pslot or ""),
                    "platform": str(platform or ""), "caption": str(caption or ""),
                    "hashtags": str(hashtags or ""), "media": str(media or ""),
                })
    except Exception as e:
        log(f"Error loading posts for slot: {e}")
    return posts

def mark_posted(post_id):
    if not os.path.exists(CONTENT_FILE): return
    try:
        wb = load_workbook(CONTENT_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value or "") == post_id:
                row[6].value = "POSTED"
                row[10].value = str(row[10].value or "") + f" | Posted: {datetime.now().strftime('%H:%M')}"
        wb.save(CONTENT_FILE)
    except Exception as e:
        log(f"Warning: mark posted failed — {e}")

# ═══════════════════════════════════════════════════════════════
# STEP 1 — NEWS MONITOR
# ═══════════════════════════════════════════════════════════════

def step_news_monitor():
    log("── STEP 1: News Monitor ──────────────────────────────")
    ensure_headlines_file()
    all_articles = []

    for feed in RSS_FEEDS:
        try:
            parsed = feedparser.parse(feed["url"])
            count  = 0
            for entry in parsed.entries[:MAX_PER_FEED]:
                title   = entry.get("title", "").strip()
                summary = re.sub(r"<[^>]+>", "", entry.get("summary", "")).strip()[:300]
                if title and is_relevant(title, summary):
                    score = sum(1 for kw in RELEVANT_KEYWORDS if kw in (title+" "+summary).lower())
                    all_articles.append({
                        "source": feed["name"], "region": feed["region"],
                        "title": title, "summary": summary, "score": score,
                    })
                    count += 1
            log(f"  {feed['name']}: {count} relevant articles")
            time.sleep(0.5)
        except Exception as e:
            log(f"  {feed['name']}: failed — {e}")

    if not all_articles:
        log("  No headlines fetched today")
        return 0

    # Save to Excel
    try:
        wb = load_workbook(HEADLINES_FILE)
        ws = wb.active
        today = datetime.now().strftime("%Y-%m-%d %H:%M")
        for a in sorted(all_articles, key=lambda x: -x["score"])[:20]:
            rn   = ws.max_row + 1
            data = [today, a["source"], a["region"], a["title"], a["summary"], a["score"]]
            for ci, val in enumerate(data, 1):
                cell = ws.cell(row=rn, column=ci, value=val)
                cell.font      = Font(name="Calibri", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=(ci in [4,5]))
            ws.row_dimensions[rn].height = 40
        wb.save(HEADLINES_FILE)
        log(f"  Saved {len(all_articles)} headlines to BIH_Marketing_Headlines.xlsx")
    except Exception as e:
        log(f"  Warning: could not save headlines — {e}")

    return len(all_articles)

# ═══════════════════════════════════════════════════════════════
# STEP 2 — CONTENT GENERATOR (Claude)
# ═══════════════════════════════════════════════════════════════

def step_content_generator(claude):
    log("── STEP 2: Content Generator (Claude) ───────────────")

    pending = get_pending_count()
    log(f"  Pending posts: {pending} / {BUFFER_TARGET} target")

    if pending >= BUFFER_TARGET:
        log("  Buffer full — skipping generation")
        return 0

    days_needed = -(-(BUFFER_TARGET - pending) // POSTS_PER_DAY)
    headlines   = load_recent_headlines(5)
    hl_text     = "\n".join(headlines) or "Write evergreen content about business intelligence and automation."
    log(f"  Generating {days_needed} day(s) of content...")

    all_posts   = []
    target_date = date.today() + timedelta(days=1)
    days_done   = 0

    while days_done < days_needed:
        if target_date.weekday() == REST_DAY:
            target_date += timedelta(days=1)
            continue

        date_str = target_date.strftime("%A, %B %d, %Y")

        # Pick one topic for the day
        try:
            resp  = claude.messages.create(
                model="claude-opus-4-6", max_tokens=120,
                messages=[{"role": "user", "content": f"""Marketing strategist for BIH (Caribbean AI & business intelligence consultancy).
Today: {date_str}. News:\n{hl_text}
Choose ONE compelling topic for today's 3 social posts, relevant to Caribbean CEOs/CFOs/Ops Managers.
ONE sentence only. No preamble."""}]
            )
            topic = resp.content[0].text.strip()
        except:
            topic = "How real-time business intelligence gives Caribbean companies a competitive edge."
        log(f"  {date_str} → Topic: {topic[:70]}")

        time.sleep(0.3)

        # Generate 3 platform posts
        for slot in ["morning", "midday", "evening"]:
            for platform, rules in PLATFORM_CONTENT.items():
                try:
                    resp    = claude.messages.create(
                        model="claude-opus-4-6", max_tokens=700,
                        messages=[{"role": "user", "content": f"""Senior copywriter for Business Intelligence Holdings (BIH).
BIH: Caribbean AI & business intelligence consultancy. Builds real-time dashboards, automates operations, implements AI for Caribbean businesses.
Founded by Omaro Hutchinson. Website: bihja.io. Book: {CALENDLY_LINK}

TOPIC: {topic}
PLATFORM: {platform}
SLOT: {slot}
VOICE: {rules['persona']}
MAX CHARS: {rules['max_chars']}
STYLE: {rules['style']}
CTA: {rules['cta']}
{"HASHTAGS (append at end): " + rules['hashtags'] if rules['hashtags'] else "NO HASHTAGS."}

RULES: Strong hook first line. Caribbean voice, not US tech cliche. No filler phrases. 1-2 sentences per paragraph.
Return ONLY the caption. No labels, no preamble."""}]
                    )
                    caption  = resp.content[0].text.strip()
                    if rules["hashtags"] and rules["hashtags"] in caption:
                        caption = caption.replace(rules["hashtags"], "").strip()
                    all_posts.append({
                        "post_date": target_date.strftime("%Y-%m-%d"),
                        "slot": slot, "platform": platform,
                        "persona": rules["persona"], "caption": caption,
                        "hashtags": rules["hashtags"],
                        "notes": f"Topic: {topic[:80]} | Claude-generated | Style: {get_day_style(target_date)['name']}",
                    })
                except Exception as e:
                    log(f"  Warning: {platform} {slot} failed — {e}")
                time.sleep(0.3)

        days_done   += 1
        target_date += timedelta(days=1)

    if all_posts:
        save_posts(all_posts)
        log(f"  Generated {len(all_posts)} posts")
    return len(all_posts)

# ═══════════════════════════════════════════════════════════════
# STEP 3 — MEDIA GENERATOR (Imagen 3 + Veo 2)
# ═══════════════════════════════════════════════════════════════

def engineer_prompt(claude, platform, topic, style, scene, specs, is_video=False):
    kind = "video generation (Veo 2)" if is_video else "image generation (Imagen 3)"
    dims = specs["video_dims"] if is_video else specs["image_dims"]
    ar   = specs["video_aspect"] if is_video else specs["image_aspect"]
    tech = style["video_technical"] if is_video else style["image_technical"]
    dur  = "10-15 seconds. " if is_video else ""

    prompt = f"""World-class creative director writing a {kind} prompt.
Brand: Business Intelligence Holdings (BIH) — Caribbean AI & business consultancy.
Brand colours: electric blue #1A3FD4, gold #F5C400, deep black background.
Platform: {platform}. Dimensions: {dims} ({ar}). {dur}Post topic: {topic}.
Visual style: {style['name']} — {style['mood']}
Base scene: {scene}
Technical: {tech}

Rules: Perfect {ar} vertical composition. Cinematic precision. No text, words or logos anywhere.
Caribbean context where natural. Blue and gold as lighting accents. Feel EXPENSIVE — not generic AI.
Return ONLY the prompt. Under 350 words."""

    try:
        resp = claude.messages.create(
            model="claude-opus-4-6", max_tokens=450,
            messages=[{"role": "user", "content": prompt}]
        )
        return resp.content[0].text.strip()
    except Exception as e:
        return f"{scene}. {tech}. {ar} vertical. No text, no logos."

def step_media_generator(claude, gemini, slot):
    log("── STEP 3: Media Generator ───────────────────────────")
    if not GEMINI_OK:
        log("  Skipping — google-genai not installed")
        return

    ensure_dirs()
    style = get_day_style()
    log(f"  Today's style: {style['name']}")

    posts = get_pending_posts_for_slot(slot)
    if not posts:
        log(f"  No posts found for {slot} slot")
        return

    is_video = (slot == "midday")
    log(f"  Generating {'VIDEO' if is_video else 'IMAGE'} for {len(posts)} post(s)...")

    for post in posts:
        platform = post["platform"]
        topic    = post["caption"][:120]
        specs    = PLATFORM_SPECS.get(platform, PLATFORM_SPECS["Instagram"])
        scenes   = style["video_scenes"] if is_video else style["image_scenes"]
        scene    = scenes.get(platform, scenes.get("Instagram", ""))

        log(f"  {platform} — engineering prompt...")
        prompt = engineer_prompt(claude, platform, topic, style, scene, specs, is_video)

        ts       = datetime.now().strftime("%Y%m%d_%H%M")
        ext      = "mp4" if is_video else "png"
        ar       = specs["video_aspect"] if is_video else specs["image_aspect"]
        fname    = f"BIH_{ts}_{slot}_{platform.replace(' ','_')}.{ext}"
        out_dir  = VIDEOS_DIR if is_video else IMAGES_DIR
        path     = os.path.join(out_dir, fname)

        try:
            if is_video:
                # Veo 2
                operation = gemini.models.generate_videos(
                    model="veo-2.0-generate-001",
                    prompt=prompt,
                    config=types.GenerateVideosConfig(
                        aspect_ratio=ar, duration_seconds=12,
                        number_of_videos=1, output_mime_type="video/mp4",
                    )
                )
                log(f"  Veo 2 generating — polling (2-5 min)...")
                waited, max_wait = 0, 360
                while not operation.done and waited < max_wait:
                    time.sleep(15); waited += 15
                    operation = gemini.operations.get(operation.name)
                if operation.done and operation.response:
                    vids = operation.response.generated_videos
                    if vids:
                        with open(path, "wb") as f: f.write(vids[0].video.video_bytes)
                        log(f"  ✅ Video saved: {fname}")
                        update_media_path(post["id"], path, "VIDEO")
            else:
                # Imagen 3
                response = gemini.models.generate_images(
                    model="imagen-3.0-generate-002",
                    prompt=prompt,
                    config=types.GenerateImagesConfig(
                        number_of_images=1, output_mime_type="image/png", aspect_ratio=ar,
                    )
                )
                if response.generated_images:
                    with open(path, "wb") as f: f.write(response.generated_images[0].image.image_bytes)
                    log(f"  ✅ Image saved: {fname}")
                    update_media_path(post["id"], path, "IMAGE")
                else:
                    log(f"  ❌ Imagen 3 returned no image")
        except Exception as e:
            log(f"  ❌ Media generation failed ({platform}): {e}")

        time.sleep(2)

# ═══════════════════════════════════════════════════════════════
# STEP 4 — POSTING (Instagram + LinkedIn + WhatsApp log)
# ═══════════════════════════════════════════════════════════════

def post_to_instagram(post):
    """Post image or video to Instagram via Meta Graph API."""
    caption  = post["caption"]
    hashtags = post["hashtags"]
    full_cap = f"{caption}\n\n{hashtags}".strip() if hashtags else caption
    media    = post.get("media", "")

    if "YOUR_" in IG_ACCESS_TOKEN:
        log("  IG: token not configured — skipping")
        return False

    try:
        # For image posts
        media_url = ""  # In production: upload image to a public URL first
        # Step 1: Create media container
        r = requests.post(
            f"https://graph.instagram.com/v18.0/{IG_ACCOUNT_ID}/media",
            params={
                "image_url":   media_url if media_url else None,
                "caption":     full_cap,
                "access_token": IG_ACCESS_TOKEN,
            }
        )
        data = r.json()
        if "id" not in data:
            log(f"  IG: container failed — {data.get('error',{}).get('message','unknown')}")
            return False

        container_id = data["id"]
        time.sleep(3)  # wait for container to process

        # Step 2: Publish
        r2 = requests.post(
            f"https://graph.instagram.com/v18.0/{IG_ACCOUNT_ID}/media_publish",
            params={"creation_id": container_id, "access_token": IG_ACCESS_TOKEN}
        )
        result = r2.json()
        if "id" in result:
            log(f"  ✅ Instagram posted (ID: {result['id']})")
            return True
        else:
            log(f"  IG publish failed — {result.get('error',{}).get('message','unknown')}")
            return False
    except Exception as e:
        log(f"  IG error: {e}")
        return False

def post_to_linkedin(post):
    """Post to LinkedIn via LinkedIn API."""
    caption  = post["caption"]
    hashtags = post["hashtags"]
    full_cap = f"{caption}\n\n{hashtags}".strip() if hashtags else caption

    if "YOUR_" in LI_ACCESS_TOKEN:
        log("  LI: token not configured — skipping")
        return False

    try:
        payload = {
            "author":          LI_PERSON_URN,
            "lifecycleState":  "PUBLISHED",
            "specificContent": {
                "com.linkedin.ugc.ShareContent": {
                    "shareCommentary":  {"text": full_cap},
                    "shareMediaCategory": "NONE",
                }
            },
            "visibility": {"com.linkedin.ugc.MemberNetworkVisibility": "PUBLIC"},
        }
        r = requests.post(
            "https://api.linkedin.com/v2/ugcPosts",
            headers={
                "Authorization":  f"Bearer {LI_ACCESS_TOKEN}",
                "Content-Type":   "application/json",
                "X-Restli-Protocol-Version": "2.0.0",
            },
            json=payload
        )
        if r.status_code in [200, 201]:
            log(f"  ✅ LinkedIn posted")
            return True
        else:
            log(f"  LI failed ({r.status_code}): {r.text[:100]}")
            return False
    except Exception as e:
        log(f"  LI error: {e}")
        return False

def post_whatsapp_log(post):
    """WhatsApp Status cannot be posted via API — log for manual posting."""
    log(f"  WhatsApp: caption saved to queue — post manually as Status")
    log(f"  Caption preview: {post['caption'][:80]}...")
    return True

def step_post(slot):
    log(f"── STEP 4: Posting ({slot}) ──────────────────────────")
    posts = get_pending_posts_for_slot(slot)

    if not posts:
        log(f"  No pending posts for {slot} slot today")
        return

    log(f"  Found {len(posts)} post(s) to publish")
    for post in posts:
        platform = post["platform"]
        log(f"\n  → {platform}")

        success = False
        if platform == "Instagram":
            success = post_to_instagram(post)
        elif platform == "LinkedIn":
            success = post_to_linkedin(post)
        elif platform == "WhatsApp":
            success = post_whatsapp_log(post)

        if success:
            mark_posted(post["id"])
        time.sleep(2)

# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════

def main():
    log("=" * 58)
    log(f"BIH Marketing AI — {datetime.now().strftime('%A %d %B %Y %H:%M')}")
    log("=" * 58)

    if date.today().weekday() == REST_DAY:
        log("Sunday — rest day. Exiting.")
        return

    ensure_dirs()
    ensure_content_file()
    ensure_headlines_file()

    slot = get_slot()
    log(f"Running slot: {slot.upper()}")

    # Init API clients
    claude = None
    gemini = None

    if ANTHROPIC_API_KEY != "YOUR_ANTHROPIC_API_KEY":
        claude = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    else:
        log("WARNING: ANTHROPIC_API_KEY not set — content generation will be skipped")

    if GEMINI_OK and GEMINI_API_KEY != "YOUR_GEMINI_API_KEY":
        gemini = genai.Client(api_key=GEMINI_API_KEY)
    else:
        log("WARNING: Gemini API key not set or google-genai not installed — media generation skipped")

    # Morning slot runs full pipeline
    if slot == "morning":
        step_news_monitor()
        if claude:
            step_content_generator(claude)
        if gemini and claude:
            step_media_generator(claude, gemini, slot)
        step_post(slot)

    # Midday and Evening just generate media + post
    elif slot in ["midday", "evening"]:
        if gemini and claude:
            step_media_generator(claude, gemini, slot)
        step_post(slot)

    log("=" * 58)
    log("BIH Marketing AI — Complete")
    log("=" * 58)

if __name__ == "__main__":
    main()
